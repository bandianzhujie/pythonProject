# -*- coding: utf-8 -*-
import openpyxl


def read_Excel(path, sheet_name, row, *col):
    # 默认从第二行开始,因为很多表都有表头
    if row == '':
        row = 2
    else:
        row = int(row)
    workbook = openpyxl.load_workbook(path)
    # 默认打开当前激活的工作表
    if sheet_name == "":
        sheet0 = workbook.active  # 获取当前激活的工作表
    else:
        sheet0 = workbook[sheet_name]  # 如果制定了工作表，就打开指定的工作表
    highest = sheet0.max_row
    case_list = {}
    # title 所在列，对比的那一列，假设A列
    title = col[0]
    for i in range(row, highest + 1):  # 遍历行
        value_list = []
        if sheet0[title + str(i)].value is None:  # 如果A5是空的，pass
            pass
        else:
            v1 = sheet0[title + str(i)].value  # 忽略大小写和前后空格
            # 除去 title的其他列
            for j in range(1, len(col)):
                v2 = sheet0[col[j] + str(i)].value
                value_list.append(v2)
            case_list[v1] = value_list
    print (case_list)
    return case_list


def write_Excel(my_dict, path, sheet_name, row, *col):
    # 将处理好的数据再次写入excel
    if row == "":
        row = 2
    else:
        row = int(row)
    workbook = openpyxl.load_workbook(path)
    if sheet_name == "":
        sheet0 = workbook.active  # 获取当前激活的工作表
    else:
        sheet0 = workbook[sheet_name]
    highest = sheet0.max_row
    # case title 所在列
    title = col[0]
    for i in range(row, highest + 1):
        if sheet0[title + str(i)].value is not None:
            v1 = sheet0[title + str(i)].value
            for key in my_dict:
                if key == v1:
                    for j in range(1, len(col)):
                        v2 = sheet0[col[j] + str(i)]
                        v2.value = my_dict[key][j - 1]
        else:
            pass

    workbook.save(path)


def sumif_Excel(path, sheet_name, row, *col):
    # 去重求和列数据
    if row == '':
        row = 2
    else:
        row = int(row)
    workbook = openpyxl.load_workbook(path)
    # 默认打开当前激活的工作表
    if sheet_name == "":
        sheet0 = workbook.active  # 获取当前激活的工作表
    else:
        sheet0 = workbook[sheet_name]  # 如果制定了工作表，就打开指定的工作表
    highest = sheet0.max_row
    case_list = {}
    # title 所在列，对比的那一列，假设A列
    title = col[0]
    for i in range(row, highest + 1):  # 遍历行
        value_list = []
        if sheet0[title + str(i)].value is None:  # 如果A5是空的，pass
            pass
        else:
            v1 = sheet0[title + str(i)].value
            v2 = sheet0[col[1] + str(i)].value
            if v1 in case_list.keys():
                value_list.append(v2 + case_list[v1][0])
            else:
                value_list.append(v2)
            case_list[v1] = value_list

    print (case_list)
    return case_list


def process(r1, r2):
    # 对比处理两次读取的内容，然后更新r2的内容
    print('Processing...')
    for key in r1:
        if key in r2:
            length = len(r1[key])
            if length > 0:
                for i in range(0, len(r1[key])):
                    # 如果想要不想覆盖原有的数值，可以取消注释，然后删除下面那行
                    if r2[key][i] is None:
                        r2[key][i] = r1[key][i]
                    # r2[key][i] = r1[key][i]
        else:
            pass

    return r2


def manual():
    info1 = raw_input('Read from fileName,[sheetName],[row],columns:\n')
    file1, sheet_name1, row1, list1 = info1.split(',')

    info2 = raw_input('Write into fileName,[sheetName],[row],columns:\n')
    file2, sheet_name2, row2, list2 = info2.split(',')

    r1 = read_Excel(file1, sheet_name1, row1, *list1)
    r2 = read_Excel(file2, sheet_name2, row2, *list2)

    r3 = process(r1, r2)
    write_Excel(r3, file2, sheet_name2, row2, *list2)

    print('Done.')


def manual1():
    info1 = raw_input('Read from fileName,[sheetName],[row],columns:\n')
    file1, sheet_name1, row1, list1 = info1.split(',')

    info2 = raw_input('Write into fileName,[sheetName],[row],columns:\n')
    file2, sheet_name2, row2, list2 = info2.split(',')

    r1 = sumif_Excel(file1, sheet_name1, row1, *list1)
    r2 = read_Excel(file2, sheet_name2, row2, *list2)

    r3 = process(r1, r2)
    write_Excel(r3, file2, sheet_name2, row2, *list2)

    print('Done.')


if __name__ == "__main__":
    info = raw_input('Do you want to sumif ? input 1 or 0 \n')
    if info == '0':
        print ('.....')
        manual()
        print ('.....')
    elif info == '1':
        manual1()
    else:
        print ('Please input right arg!')
        pass


